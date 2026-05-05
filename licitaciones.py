#!/usr/bin/env python3
"""
Checklist de Carpetas de Licitación
Autor: Sistema de gestión documental
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import os
import json
import re
from datetime import datetime
import openpyxl
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, HRFlowable, KeepTogether)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY

# ── Colores institucionales ──────────────────────────────────────────────────
C_BG       = "#F0F4F8"
C_SIDEBAR  = "#1A3A5C"
C_HEADER   = "#1A3A5C"
C_ACCENT   = "#2E86C1"
C_SUCCESS  = "#27AE60"
C_WARNING  = "#E67E22"
C_DANGER   = "#C0392B"
C_NA       = "#7F8C8D"
C_WEB      = "#8E44AD"
C_WEB_OK   = "#9B59B6"
C_WHITE    = "#FFFFFF"
C_LIGHT    = "#ECF0F1"
C_TEXT     = "#2C3E50"
C_MUTED    = "#95A5A6"

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "licitaciones.db")

# ═══════════════════════════════════════════════════════════════════════════════
# BASE DE DATOS
# ═══════════════════════════════════════════════════════════════════════════════
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS licitaciones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo TEXT UNIQUE NOT NULL,
        nombre TEXT NOT NULL,
        tipo TEXT NOT NULL DEFAULT 'Publica',
        fecha_creacion TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS propuestas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        licitacion_id INTEGER,
        codigo TEXT NOT NULL,
        FOREIGN KEY(licitacion_id) REFERENCES licitaciones(id),
        UNIQUE(licitacion_id, codigo)
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS documentos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        tipo TEXT NOT NULL,
        solo_privada INTEGER DEFAULT 0,
        control_web INTEGER DEFAULT 0,
        orden INTEGER DEFAULT 0
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS checklist (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        licitacion_id INTEGER,
        propuesta_id INTEGER,
        documento_id INTEGER,
        estado TEXT DEFAULT 'No',
        web_marcado INTEGER DEFAULT 0,
        numero TEXT DEFAULT '',
        fecha_doc TEXT DEFAULT '',
        FOREIGN KEY(licitacion_id) REFERENCES licitaciones(id),
        FOREIGN KEY(propuesta_id) REFERENCES propuestas(id),
        FOREIGN KEY(documento_id) REFERENCES documentos(id),
        UNIQUE(licitacion_id, propuesta_id, documento_id)
    )""")
    try:
        c.execute("ALTER TABLE checklist ADD COLUMN web_marcado INTEGER DEFAULT 0")
    except:
        pass
    try:
        c.execute("ALTER TABLE documentos ADD COLUMN control_web INTEGER DEFAULT 0")
    except:
        pass
    conn.commit()
    conn.close()

def get_conn():
    return sqlite3.connect(DB_PATH)

# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def es_solo_privada(nombre_doc):
    return "(solo en caso de ser licitación privada)" in nombre_doc.lower()

def nombre_limpio(nombre_doc):
    return re.sub(r'\s*\(solo en caso de ser licitación privada\)\s*', '', nombre_doc, flags=re.IGNORECASE).strip()

def es_resolucion_oficio(nombre_doc):
    nl = nombre_limpio(nombre_doc).lower()
    keywords = ["resolución", "resolucion", "oficio", "decreto"]
    return any(k in nl for k in keywords)

def init_checklist_licitacion(conn, lic_id):
    c = conn.cursor()
    docs = c.execute("SELECT id FROM documentos WHERE tipo='licitacion'").fetchall()
    for (doc_id,) in docs:
        c.execute("""INSERT OR IGNORE INTO checklist(licitacion_id, propuesta_id, documento_id, estado, web_marcado)
                     VALUES(?,NULL,?,'No',0)""", (lic_id, doc_id))
    conn.commit()

def init_checklist_propuesta(conn, lic_id, prop_id):
    c = conn.cursor()
    docs = c.execute("SELECT id FROM documentos WHERE tipo='propuesta'").fetchall()
    for (doc_id,) in docs:
        c.execute("""INSERT OR IGNORE INTO checklist(licitacion_id, propuesta_id, documento_id, estado, web_marcado)
                     VALUES(?,?,?,'No',0)""", (lic_id, prop_id, doc_id))
    conn.commit()

# ═══════════════════════════════════════════════════════════════════════════════
# APLICACIÓN PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Checklist Carpetas de Licitación")
        self.geometry("1280x800")
        self.minsize(1100, 700)
        self.configure(bg=C_BG)
        init_db()
        self._build_ui()
        self._refresh_sidebar()

    def _build_ui(self):
        self.sidebar = tk.Frame(self, bg=C_SIDEBAR, width=260)
        self.sidebar.pack(side=tk.LEFT, fill=tk.Y)
        self.sidebar.pack_propagate(False)

        tk.Label(self.sidebar, text="📁 Licitaciones", bg=C_SIDEBAR, fg=C_WHITE,
                 font=("Segoe UI", 13, "bold"), pady=16).pack(fill=tk.X)

        HrLine(self.sidebar, C_ACCENT).pack(fill=tk.X, padx=10)

        btn_frame = tk.Frame(self.sidebar, bg=C_SIDEBAR)
        btn_frame.pack(fill=tk.X, pady=8, padx=10)
        SideBtn(btn_frame, "➕  Nueva Licitación", self._dlg_nueva_licitacion).pack(fill=tk.X, pady=2)
        SideBtn(btn_frame, "📥  Cargar Licitaciones Excel", self._import_licitaciones).pack(fill=tk.X, pady=2)
        SideBtn(btn_frame, "📄  Cargar Documentos Excel", self._import_documentos).pack(fill=tk.X, pady=2)

        HrLine(self.sidebar, C_ACCENT).pack(fill=tk.X, padx=10, pady=4)

        search_frame = tk.Frame(self.sidebar, bg=C_SIDEBAR)
        search_frame.pack(fill=tk.X, padx=10, pady=(0,6))
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self._refresh_sidebar())
        tk.Entry(search_frame, textvariable=self.search_var, bg="#2C5282", fg=C_WHITE,
                 insertbackground=C_WHITE, relief=tk.FLAT, font=("Segoe UI", 9),
                 ).pack(fill=tk.X, ipady=5, padx=2)

        list_frame = tk.Frame(self.sidebar, bg=C_SIDEBAR)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=6)
        sb = ttk.Scrollbar(list_frame, orient=tk.VERTICAL)
        self.lic_listbox = tk.Listbox(list_frame, bg="#162D46", fg=C_WHITE,
                                       selectbackground=C_ACCENT, selectforeground=C_WHITE,
                                       relief=tk.FLAT, font=("Segoe UI", 9),
                                       yscrollcommand=sb.set, activestyle="none",
                                       highlightthickness=0, borderwidth=0)
        sb.config(command=self.lic_listbox.yview)
        self.lic_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.lic_listbox.bind("<<ListboxSelect>>", self._on_lic_select)
        self._lic_ids = []

        self.content = tk.Frame(self, bg=C_BG)
        self.content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._show_welcome()

    def _show_welcome(self):
        for w in self.content.winfo_children():
            w.destroy()
        fr = tk.Frame(self.content, bg=C_BG)
        fr.place(relx=0.5, rely=0.45, anchor=tk.CENTER)
        tk.Label(fr, text="📋", font=("Segoe UI", 56), bg=C_BG).pack()
        tk.Label(fr, text="Checklist Carpetas de Licitación",
                 font=("Segoe UI", 18, "bold"), bg=C_BG, fg=C_HEADER).pack(pady=(8,4))
        tk.Label(fr, text="Seleccione una licitación del panel izquierdo\no cree una nueva para comenzar.",
                 font=("Segoe UI", 11), bg=C_BG, fg=C_MUTED, justify=tk.CENTER).pack()

    def _refresh_sidebar(self, *_):
        conn = get_conn()
        q = self.search_var.get().strip().lower()
        rows = conn.execute(
            "SELECT id, codigo, nombre, tipo FROM licitaciones ORDER BY codigo"
        ).fetchall()
        conn.close()
        self.lic_listbox.delete(0, tk.END)
        self._lic_ids = []
        for row in rows:
            lid, cod, nom, tipo = row
            label = f"{cod} — {nom}"
            if q and q not in label.lower():
                continue
            icon = "🔒" if tipo == "Privada" else "🌐"
            self.lic_listbox.insert(tk.END, f"  {icon} {cod}")
            self._lic_ids.append(lid)

    def _on_lic_select(self, _=None):
        sel = self.lic_listbox.curselection()
        if not sel:
            return
        lic_id = self._lic_ids[sel[0]]
        self._open_licitacion(lic_id)

    def _dlg_nueva_licitacion(self):
        dlg = tk.Toplevel(self)
        dlg.title("Nueva Licitación")
        dlg.geometry("420x280")
        dlg.configure(bg=C_BG)
        dlg.grab_set()
        dlg.resizable(False, False)

        tk.Label(dlg, text="Nueva Licitación", font=("Segoe UI", 13, "bold"),
                 bg=C_BG, fg=C_HEADER).pack(pady=(18,8))

        form = tk.Frame(dlg, bg=C_BG)
        form.pack(padx=30, fill=tk.X)

        def field(label, row):
            tk.Label(form, text=label, bg=C_BG, fg=C_TEXT,
                     font=("Segoe UI", 9, "bold"), anchor="w").grid(row=row, column=0, sticky="w", pady=4)
            e = tk.Entry(form, font=("Segoe UI", 10), relief=tk.SOLID, bd=1)
            e.grid(row=row, column=1, sticky="ew", padx=(8,0), ipady=4)
            return e

        form.columnconfigure(1, weight=1)
        e_cod = field("Código:", 0)
        e_nom = field("Nombre:", 1)

        tk.Label(form, text="Tipo:", bg=C_BG, fg=C_TEXT,
                 font=("Segoe UI", 9, "bold"), anchor="w").grid(row=2, column=0, sticky="w", pady=4)
        tipo_var = tk.StringVar(value="Publica")
        tipo_cb = ttk.Combobox(form, textvariable=tipo_var, values=["Publica", "Privada"],
                               state="readonly", font=("Segoe UI", 10))
        tipo_cb.grid(row=2, column=1, sticky="ew", padx=(8,0), ipady=2)

        def save():
            cod = e_cod.get().strip()
            nom = e_nom.get().strip()
            if not cod or not nom:
                messagebox.showwarning("Faltan datos", "Código y Nombre son obligatorios.", parent=dlg)
                return
            conn = get_conn()
            try:
                conn.execute(
                    "INSERT INTO licitaciones(codigo,nombre,tipo,fecha_creacion) VALUES(?,?,?,?)",
                    (cod, nom, tipo_var.get(), datetime.now().strftime("%Y-%m-%d"))
                )
                conn.commit()
                lic_id = conn.execute("SELECT id FROM licitaciones WHERE codigo=?", (cod,)).fetchone()[0]
                init_checklist_licitacion(conn, lic_id)
            except sqlite3.IntegrityError:
                messagebox.showerror("Duplicado", f"Ya existe la licitación '{cod}'.", parent=dlg)
                conn.close()
                return
            conn.close()
            dlg.destroy()
            self._refresh_sidebar()
            self._open_licitacion(lic_id)

        tk.Button(dlg, text="Crear Licitación", command=save,
                  bg=C_ACCENT, fg=C_WHITE, font=("Segoe UI", 10, "bold"),
                  relief=tk.FLAT, cursor="hand2", pady=8).pack(pady=18, padx=30, fill=tk.X)

    def _import_licitaciones(self):
        path = filedialog.askopenfilename(
            title="Seleccionar Excel de Licitaciones",
            filetypes=[("Excel", "*.xlsx *.xls")]
        )
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            conn = get_conn()
            ok = 0
            skip = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                cod = str(row[0]).strip()
                nom = str(row[1]).strip() if len(row) > 1 and row[1] else cod
                tipo_raw = str(row[2]).strip() if len(row) > 2 and row[2] else "Publica"
                tipo = "Privada" if "priv" in tipo_raw.lower() else "Publica"
                try:
                    conn.execute(
                        "INSERT INTO licitaciones(codigo,nombre,tipo,fecha_creacion) VALUES(?,?,?,?)",
                        (cod, nom, tipo, datetime.now().strftime("%Y-%m-%d"))
                    )
                    conn.commit()
                    lic_id = conn.execute("SELECT id FROM licitaciones WHERE codigo=?", (cod,)).fetchone()[0]
                    init_checklist_licitacion(conn, lic_id)
                    ok += 1
                except sqlite3.IntegrityError:
                    skip += 1
            conn.close()
            messagebox.showinfo("Importación completada",
                                f"✅ {ok} licitaciones importadas.\n⚠️ {skip} ya existían (omitidas).")
            self._refresh_sidebar()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo:\n{e}")

    def _import_documentos(self):
        path = filedialog.askopenfilename(
            title="Seleccionar Excel de Documentos",
            filetypes=[("Excel", "*.xlsx *.xls")]
        )
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            conn = get_conn()
            conn.execute("DELETE FROM checklist")
            conn.execute("DELETE FROM documentos")
            conn.commit()
            orden = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                
                doc_lic = str(row[0]).strip() if len(row) > 0 and row[0] else None
                control_web_lic = 0
                if len(row) > 1 and row[1]:
                    valor = str(row[1]).strip().upper()
                    control_web_lic = 1 if valor in ['SI', 'S', 'YES', 'Y', '1', 'WEB'] else 0
                
                doc_prop = str(row[2]).strip() if len(row) > 2 and row[2] else None
                
                if doc_lic:
                    solo = 1 if es_solo_privada(doc_lic) else 0
                    conn.execute("INSERT INTO documentos(nombre,tipo,solo_privada,control_web,orden) VALUES(?,?,?,?,?)",
                                 (doc_lic, "licitacion", solo, control_web_lic, orden))
                if doc_prop:
                    solo = 1 if es_solo_privada(doc_prop) else 0
                    conn.execute("INSERT INTO documentos(nombre,tipo,solo_privada,control_web,orden) VALUES(?,?,?,?,?)",
                                 (doc_prop, "propuesta", solo, 0, orden))
                orden += 1
            conn.commit()
            
            lics = conn.execute("SELECT id FROM licitaciones").fetchall()
            for (lid,) in lics:
                init_checklist_licitacion(conn, lid)
                props = conn.execute("SELECT id FROM propuestas WHERE licitacion_id=?", (lid,)).fetchall()
                for (pid,) in props:
                    init_checklist_propuesta(conn, lid, pid)
            conn.close()
            messagebox.showinfo("Documentos cargados",
                                "✅ Lista de documentos actualizada correctamente.\n\n"
                                "Estructura del Excel:\n"
                                "Col A: Documentos Licitación\n"
                                "Col B: Control Web (SI/NO)\n"
                                "Col C: Documentos Propuesta")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo:\n{e}")

    def _open_licitacion(self, lic_id):
        for w in self.content.winfo_children():
            w.destroy()
        LicitacionView(self.content, lic_id, self._refresh_sidebar).pack(fill=tk.BOTH, expand=True)
    
    def _show_welcome_again(self):
        self._show_welcome()


# ═══════════════════════════════════════════════════════════════════════════════
# VISTA DE UNA LICITACIÓN
# ═══════════════════════════════════════════════════════════════════════════════
class LicitacionView(tk.Frame):
    def __init__(self, parent, lic_id, refresh_cb):
        super().__init__(parent, bg=C_BG)
        self.lic_id = lic_id
        self.refresh_cb = refresh_cb
        self._load_data()
        self.solo_falta_lic = tk.BooleanVar(value=False)
        self.solo_falta_todos = tk.BooleanVar(value=False)
        self._build()

    def _load_data(self):
        conn = get_conn()
        row = conn.execute(
            "SELECT codigo, nombre, tipo FROM licitaciones WHERE id=?", (self.lic_id,)
        ).fetchone()
        self.codigo, self.nombre, self.tipo = row
        conn.close()

    def _build(self):
        hdr = tk.Frame(self, bg=C_HEADER)
        hdr.pack(fill=tk.X)

        tk.Label(hdr, text=f"  {self.codigo}  —  {self.nombre}",
                 font=("Segoe UI", 13, "bold"), bg=C_HEADER, fg=C_WHITE,
                 pady=12).pack(side=tk.LEFT)

        tipo_color = C_DANGER if self.tipo == "Privada" else C_SUCCESS
        tk.Label(hdr, text=f"  {self.tipo}  ", font=("Segoe UI", 9, "bold"),
                 bg=tipo_color, fg=C_WHITE, pady=4, padx=6).pack(side=tk.LEFT, padx=8, pady=10)

        btn_hdr = tk.Frame(hdr, bg=C_HEADER)
        btn_hdr.pack(side=tk.RIGHT, padx=10)
        HdrBtn(btn_hdr, "📄  Acta de Cierre PDF", self._generar_pdf).pack(side=tk.RIGHT, padx=4, pady=8)
        HdrBtn(btn_hdr, "🗑  Eliminar", self._eliminar).pack(side=tk.RIGHT, padx=4, pady=8)

        style = ttk.Style()
        style.configure("Custom.TNotebook", background=C_BG, borderwidth=0)
        style.configure("Custom.TNotebook.Tab", font=("Segoe UI", 10),
                         padding=[16, 8], background=C_LIGHT, foreground=C_TEXT)
        style.map("Custom.TNotebook.Tab",
                  background=[("selected", C_WHITE)],
                  foreground=[("selected", C_ACCENT)])

        self.nb = ttk.Notebook(self, style="Custom.TNotebook")
        self.nb.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

        self.tab_lic = tk.Frame(self.nb, bg=C_BG)
        self.tab_prop = tk.Frame(self.nb, bg=C_BG)
        self.tab_todos = tk.Frame(self.nb, bg=C_BG)

        self.nb.add(self.tab_lic, text="  Checklist Licitación  ")
        self.nb.add(self.tab_prop, text="  Propuestas  ")
        self.nb.add(self.tab_todos, text="  Todos los Documentos  ")

        self._build_tab_licitacion()
        self._build_tab_propuestas()
        self._build_tab_todos()

    def _build_tab_licitacion(self):
        fr = self.tab_lic
        for w in fr.winfo_children():
            w.destroy()

        conn = get_conn()
        docs = conn.execute(
            "SELECT id, nombre, solo_privada, control_web FROM documentos WHERE tipo='licitacion' ORDER BY orden"
        ).fetchall()
        
        checks_raw = conn.execute(
            "SELECT id, estado, documento_id, numero, fecha_doc, web_marcado FROM checklist "
            "WHERE licitacion_id=? AND propuesta_id IS NULL", (self.lic_id,)
        ).fetchall()
        
        # Agrega este print para ver qué datos se están cargando
        print(f"\n=== CARGANDO DATOS PARA LICITACION ID={self.lic_id} ===")
        for row in checks_raw:
            print(f"    doc_id={row[2]}, estado={row[1]}, web_marcado={row[5]}")
            
        checks = {}
        for row in checks_raw:
            chk_id, estado, doc_id, numero, fecha_doc, web_marcado = row
            checks[doc_id] = (chk_id, estado, numero, fecha_doc, web_marcado)
        
        conn.close()

        top = tk.Frame(fr, bg=C_BG)
        top.pack(fill=tk.X, padx=20, pady=(14,8))

        chk = tk.Checkbutton(top, text="Mostrar solo pendientes",
                             variable=self.solo_falta_lic,
                             command=self._build_tab_licitacion,
                             bg=C_BG, fg=C_TEXT, font=("Segoe UI", 9),
                             activebackground=C_BG, selectcolor=C_WHITE)
        chk.pack(side=tk.LEFT)

        total_docs = [d for d in docs if not (d[2] == 1 and self.tipo == "Publica")]
        completados = 0
        na_count = 0
        for d in total_docs:
            doc_id, _, _, control_web = d
            chk_row = checks.get(doc_id, (None, "No", "", "", 0))
            estado = chk_row[1]
            web_marcado = chk_row[4]
            if estado == "NA":
                na_count += 1
            elif estado == "Si" and (web_marcado == 1 or control_web == 0):
                completados += 1
        
        efectivos = len(total_docs) - na_count
        pct = int(completados / efectivos * 100) if efectivos > 0 else 0

        prog_fr = tk.Frame(top, bg=C_BG)
        prog_fr.pack(side=tk.RIGHT)
        tk.Label(prog_fr, text=f"Avance: {completados}/{efectivos}  ({pct}%)",
                 font=("Segoe UI", 9, "bold"), bg=C_BG, fg=C_ACCENT).pack()

        canvas, inner = scrollable(fr)

        mostrar_solo_pendientes = self.solo_falta_lic.get()
        
        for doc_id, doc_nombre, solo_priv, control_web in docs:
            es_privada_doc = solo_priv == 1
            locked = es_privada_doc and self.tipo == "Publica"

            chk_row = checks.get(doc_id, (None, "No", "", "", 0))
            chk_id, estado, numero, fecha_doc, web_marcado = chk_row

            esta_completo = (estado == "Si" and (web_marcado == 1 or control_web == 0))
            es_pendiente = not (estado == "NA" or esta_completo)
            
            if mostrar_solo_pendientes and not es_pendiente:
                continue

            DocRow(inner, self.lic_id, None, doc_id, doc_nombre, estado,
                   numero, fecha_doc, web_marcado, locked, self.tipo, control_web,
                   self._build_tab_licitacion,
                   self._build_tab_todos).pack(fill=tk.X, padx=12, pady=2)

    def _build_tab_propuestas(self):
        fr = self.tab_prop
        for w in fr.winfo_children():
            w.destroy()

        top = tk.Frame(fr, bg=C_BG)
        top.pack(fill=tk.X, padx=20, pady=(14,8))

        def add_prop():
            cod = simple_input(fr, "Código de Propuesta")
            if not cod:
                return
            conn = get_conn()
            try:
                conn.execute(
                    "INSERT INTO propuestas(licitacion_id, codigo) VALUES(?,?)",
                    (self.lic_id, cod)
                )
                conn.commit()
                pid = conn.execute(
                    "SELECT id FROM propuestas WHERE licitacion_id=? AND codigo=?",
                    (self.lic_id, cod)
                ).fetchone()[0]
                init_checklist_propuesta(conn, self.lic_id, pid)
            except sqlite3.IntegrityError:
                messagebox.showwarning("Duplicado", f"La propuesta '{cod}' ya existe.")
            conn.close()
            self._build_tab_propuestas()
            self._build_tab_todos()

        def import_props():
            path = filedialog.askopenfilename(
                title="Seleccionar Excel de Propuestas",
                filetypes=[("Excel", "*.xlsx *.xls")]
            )
            if not path:
                return
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb.active
                conn = get_conn()
                ok = skip = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    cod = str(row[0]).strip()
                    try:
                        conn.execute(
                            "INSERT INTO propuestas(licitacion_id, codigo) VALUES(?,?)",
                            (self.lic_id, cod)
                        )
                        conn.commit()
                        pid = conn.execute(
                            "SELECT id FROM propuestas WHERE licitacion_id=? AND codigo=?",
                            (self.lic_id, cod)
                        ).fetchone()[0]
                        init_checklist_propuesta(conn, self.lic_id, pid)
                        ok += 1
                    except sqlite3.IntegrityError:
                        skip += 1
                conn.close()
                messagebox.showinfo("Listo", f"✅ {ok} propuestas importadas. {skip} ya existían.")
                self._build_tab_propuestas()
                self._build_tab_todos()
            except Exception as e:
                messagebox.showerror("Error", str(e))

        HdrBtn(top, "➕  Nueva Propuesta", add_prop).pack(side=tk.LEFT, padx=(0,8))
        HdrBtn(top, "📥  Cargar Propuestas Excel", import_props).pack(side=tk.LEFT)

        canvas, inner = scrollable(fr)

        conn = get_conn()
        props = conn.execute(
            "SELECT id, codigo FROM propuestas WHERE licitacion_id=? ORDER BY codigo",
            (self.lic_id,)
        ).fetchall()
        conn.close()

        for prop_id, prop_cod in props:
            PropCard(inner, self.lic_id, prop_id, prop_cod, self.tipo,
                     self._build_tab_propuestas, self._build_tab_todos).pack(
                fill=tk.X, padx=12, pady=6)

    def _build_tab_todos(self):
        fr = self.tab_todos
        for w in fr.winfo_children():
            w.destroy()

        top = tk.Frame(fr, bg=C_BG)
        top.pack(fill=tk.X, padx=20, pady=(14,8))
        
        tk.Checkbutton(top, text="Mostrar solo pendientes",
                       variable=self.solo_falta_todos,
                       command=self._build_tab_todos,
                       bg=C_BG, fg=C_TEXT, font=("Segoe UI", 9),
                       activebackground=C_BG, selectcolor=C_WHITE).pack(side=tk.LEFT)

        canvas, inner = scrollable(fr)

        conn = get_conn()
        docs_lic = conn.execute(
            "SELECT id, nombre, solo_privada, control_web FROM documentos WHERE tipo='licitacion' ORDER BY orden"
        ).fetchall()
        
        checks_lic_raw = conn.execute(
            "SELECT id, estado, documento_id, numero, fecha_doc, web_marcado FROM checklist "
            "WHERE licitacion_id=? AND propuesta_id IS NULL", (self.lic_id,)
        ).fetchall()
        
        checks_lic = {}
        for row in checks_lic_raw:
            chk_id, estado, doc_id, numero, fecha_doc, web_marcado = row
            checks_lic[doc_id] = (chk_id, estado, numero, fecha_doc, web_marcado)

        props = conn.execute(
            "SELECT id, codigo FROM propuestas WHERE licitacion_id=? ORDER BY codigo",
            (self.lic_id,)
        ).fetchall()
        conn.close()

        mostrar_solo_pendientes = self.solo_falta_todos.get()

        SectionHeader(inner, "📋  Documentos de la Licitación").pack(fill=tk.X, padx=12, pady=(8,4))
        for doc_id, doc_nombre, solo_priv, control_web in docs_lic:
            locked = (solo_priv == 1 and self.tipo == "Publica")
            chk_row = checks_lic.get(doc_id, (None, "No", "", "", 0))
            chk_id, estado, numero, fecha_doc, web_marcado = chk_row
            
            esta_completo = (estado == "Si" and (web_marcado == 1 or control_web == 0))
            es_pendiente = not (estado == "NA" or esta_completo)
            
            if mostrar_solo_pendientes and not es_pendiente:
                continue
            
            DocRow(inner, self.lic_id, None, doc_id, doc_nombre, estado,
                   numero, fecha_doc, web_marcado, locked, self.tipo, control_web,
                   self._build_tab_todos, None).pack(fill=tk.X, padx=12, pady=2)

        for prop_id, prop_cod in props:
            SectionHeader(inner, f"📁  Propuesta: {prop_cod}").pack(fill=tk.X, padx=12, pady=(12,4))
            conn2 = get_conn()
            docs_prop = conn2.execute(
                "SELECT id, nombre, solo_privada, control_web FROM documentos WHERE tipo='propuesta' ORDER BY orden"
            ).fetchall()
            
            checks_prop_raw = conn2.execute(
                "SELECT id, estado, documento_id, numero, fecha_doc, web_marcado FROM checklist "
                "WHERE licitacion_id=? AND propuesta_id=?", (self.lic_id, prop_id)
            ).fetchall()
            
            checks_prop = {}
            for row in checks_prop_raw:
                chk_id, estado, doc_id, numero, fecha_doc, web_marcado = row
                checks_prop[doc_id] = (chk_id, estado, numero, fecha_doc, web_marcado)
            
            conn2.close()
            
            for doc_id, doc_nombre, solo_priv, control_web in docs_prop:
                locked = (solo_priv == 1 and self.tipo == "Publica")
                chk_row = checks_prop.get(doc_id, (None, "No", "", "", 0))
                chk_id, estado, numero, fecha_doc, web_marcado = chk_row
                
                esta_completo = (estado == "Si" and (web_marcado == 1 or control_web == 0))
                es_pendiente = not (estado == "NA" or esta_completo)
                
                if mostrar_solo_pendientes and not es_pendiente:
                    continue
                
                DocRow(inner, self.lic_id, prop_id, doc_id, doc_nombre, estado,
                       numero, fecha_doc, web_marcado, locked, self.tipo, control_web,
                       self._build_tab_todos, None).pack(fill=tk.X, padx=12, pady=2)

    def _eliminar(self):
        if not messagebox.askyesno("Confirmar",
                                    f"¿Eliminar la licitación '{self.codigo}'?\n"
                                    "Se perderá todo el progreso."):
            return
        conn = get_conn()
        conn.execute("DELETE FROM checklist WHERE licitacion_id=?", (self.lic_id,))
        conn.execute("DELETE FROM propuestas WHERE licitacion_id=?", (self.lic_id,))
        conn.execute("DELETE FROM licitaciones WHERE id=?", (self.lic_id,))
        conn.commit()
        conn.close()
        self.refresh_cb()
        self.master._show_welcome_again()

    def _generar_pdf(self):
        # Preguntar quién firma
        firmante = simple_input(self, "Nombre del firmante")
        if not firmante:
            return  # Si no ingresa nombre, cancelar
    
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"Acta_Cierre_{self.codigo}.pdf"
        )
        if not path:
            return
        try:
            generar_acta_pdf(path, self.lic_id, self.codigo, self.nombre, self.tipo, firmante)
            messagebox.showinfo("PDF generado", f"✅ Acta de cierre guardada en:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el PDF:\n{e}")


# ═══════════════════════════════════════════════════════════════════════════════
# FILA DE DOCUMENTO - CON WEB INDEPENDIENTE
# ═══════════════════════════════════════════════════════════════════════════════
class DocRow(tk.Frame):
    ESTADOS = ["No", "Si", "NA"]
    LABELS = {"No": "✗  No", "Si": "✔  Sí", "NA": "—  N/A"}
    COLORS = {"No": C_DANGER, "Si": C_SUCCESS, "NA": C_NA}

    def __init__(self, parent, lic_id, prop_id, doc_id, doc_nombre,
                 estado, numero, fecha_doc, web_marcado, locked, tipo_lic, control_web,
                 refresh_cb, refresh_todos_cb):
        super().__init__(parent, bg=C_WHITE, relief=tk.FLAT,
                         highlightbackground="#DDE3EA", highlightthickness=1)
        self.lic_id = lic_id
        self.prop_id = prop_id
        self.doc_id = doc_id
        self.estado = estado
        self.numero = numero or ""
        self.fecha_doc = fecha_doc or ""
        self.web_marcado = web_marcado or 0
        self.locked = locked
        self.tipo_lic = tipo_lic
        self.control_web = control_web
        self.refresh_cb = refresh_cb
        self.refresh_todos_cb = refresh_todos_cb
        self.nombre_raw = doc_nombre
        self._build()

    def _get_color_fondo(self):
        requiere_web = self.control_web == 1 and not self.locked
        
        if self.locked:
            return C_LIGHT
        elif self.estado == "NA":
            return "#F8F9FA"
        elif self.estado == "Si" and (self.web_marcado == 1 or not requiere_web):
            return "#F0FFF4"
        elif self.estado == "Si" and self.web_marcado == 0 and requiere_web:
            return "#FFF8E1"
        elif self.estado == "No" and self.web_marcado == 1 and requiere_web:
            return "#FFF8E1"
        else:
            return "#FFF5F5"

    def _build(self):
        nombre_disp = nombre_limpio(self.nombre_raw)
        solo_priv = es_solo_privada(self.nombre_raw)
        es_res = es_resolucion_oficio(self.nombre_raw)
        requiere_web = self.control_web == 1 and not self.locked
        
        bg_color = self._get_color_fondo()
        self.configure(bg=bg_color)

        row = tk.Frame(self, bg=bg_color)
        row.pack(fill=tk.X, padx=10, pady=6)

        suffix = "  🔒 Solo licitación privada" if solo_priv else ""
        web_suffix = "  🌐 Requiere web" if requiere_web else ""
        lbl_text = nombre_disp + suffix + web_suffix
        tk.Label(row, text=lbl_text, font=("Segoe UI", 9),
                 bg=bg_color, fg=C_MUTED if self.locked else C_TEXT,
                 anchor="w", wraplength=450, justify=tk.LEFT).pack(side=tk.LEFT, fill=tk.X, expand=True)

        btn_frame = tk.Frame(row, bg=bg_color)
        btn_frame.pack(side=tk.RIGHT)

        if self.locked:
            tk.Label(btn_frame, text="N/A  (Pública)", font=("Segoe UI", 8),
                     bg=C_NA, fg=C_WHITE, padx=8, pady=4).pack(side=tk.RIGHT)
        else:
            for est in self.ESTADOS:
                active = (est == self.estado)
                color = self.COLORS[est] if active else "#DEE2E6"
                fg = C_WHITE if active else C_TEXT
                
                btn = tk.Button(btn_frame, text=self.LABELS[est],
                                bg=color, fg=fg,
                                font=("Segoe UI", 8, "bold" if active else "normal"),
                                relief=tk.FLAT, cursor="hand2", padx=8, pady=4,
                                command=lambda e=est: self._set_estado(e))
                btn.pack(side=tk.LEFT, padx=2)
            
            if requiere_web:
                web_color = C_WEB_OK if self.web_marcado == 1 else "#DEE2E6"
                web_fg = C_WHITE if self.web_marcado == 1 else C_TEXT
                web_btn = tk.Button(btn_frame, text="🌐  Web",
                                    bg=web_color, fg=web_fg,
                                    font=("Segoe UI", 8, "bold" if self.web_marcado == 1 else "normal"),
                                    relief=tk.FLAT, cursor="hand2", padx=8, pady=4,
                                    command=self._toggle_web)
                web_btn.pack(side=tk.LEFT, padx=2)

        if es_res and not self.locked:
            extra = tk.Frame(self, bg=bg_color)
            extra.pack(fill=tk.X, padx=10, pady=(0,6))

            tk.Label(extra, text="N°:", font=("Segoe UI", 8), bg=bg_color, fg=C_MUTED).pack(side=tk.LEFT)
            self.e_num = tk.Entry(extra, width=14, font=("Segoe UI", 8), relief=tk.SOLID, bd=1)
            self.e_num.insert(0, self.numero)
            self.e_num.pack(side=tk.LEFT, padx=(2,12), ipady=2)

            tk.Label(extra, text="Fecha:", font=("Segoe UI", 8), bg=bg_color, fg=C_MUTED).pack(side=tk.LEFT)
            self.e_fec = tk.Entry(extra, width=14, font=("Segoe UI", 8), relief=tk.SOLID, bd=1)
            self.e_fec.insert(0, self.fecha_doc)
            self.e_fec.pack(side=tk.LEFT, padx=(2,12), ipady=2)

            tk.Button(extra, text="Guardar", command=self._save_extra,
                      bg=C_ACCENT, fg=C_WHITE, font=("Segoe UI", 8),
                      relief=tk.FLAT, cursor="hand2", padx=6, pady=2).pack(side=tk.LEFT)

    def _set_estado(self, nuevo_estado):
        conn = get_conn()
        conn.execute(
            """INSERT INTO checklist(licitacion_id, propuesta_id, documento_id, estado, web_marcado)
               VALUES(?,?,?,?,?)
               ON CONFLICT(licitacion_id, propuesta_id, documento_id)
               DO UPDATE SET estado=excluded.estado, web_marcado=excluded.web_marcado""",
             (self.lic_id, self.prop_id, self.doc_id, nuevo_estado, self.web_marcado)
        )
        conn.commit()
        conn.close()
        self.estado = nuevo_estado
        if self.refresh_cb:
            self.refresh_cb()
        if self.refresh_todos_cb:
            self.refresh_todos_cb()

    def _toggle_web(self):
        nuevo_valor = 1 if self.web_marcado == 0 else 0
        conn = get_conn()
        conn.execute(
            """INSERT INTO checklist(licitacion_id, propuesta_id, documento_id, estado, web_marcado)
               VALUES(?,?,?,?,?)
               ON CONFLICT(licitacion_id, propuesta_id, documento_id)
               DO UPDATE SET web_marcado=excluded.web_marcado, estado=excluded.estado""",
            (self.lic_id, self.prop_id, self.doc_id, self.estado, nuevo_valor)
        )
        conn.commit()
        conn.close()
        self.web_marcado = nuevo_valor
        if self.refresh_cb:
            self.refresh_cb()
        if self.refresh_todos_cb:
            self.refresh_todos_cb()
    def _save_extra(self):
        num = self.e_num.get().strip()
        fec = self.e_fec.get().strip()
        conn = get_conn()
        conn.execute(
            """INSERT INTO checklist(licitacion_id, propuesta_id, documento_id, numero, fecha_doc)
               VALUES(?,?,?,?,?)
               ON CONFLICT(licitacion_id, propuesta_id, documento_id)
               DO UPDATE SET numero=excluded.numero, fecha_doc=excluded.fecha_doc""",
            (self.lic_id, self.prop_id, self.doc_id, num, fec)
        )
        conn.commit()
        conn.close()
        messagebox.showinfo("Guardado", "N° y Fecha guardados correctamente.")


# ═══════════════════════════════════════════════════════════════════════════════
# TARJETA DE PROPUESTA
# ═══════════════════════════════════════════════════════════════════════════════
class PropCard(tk.Frame):
    def __init__(self, parent, lic_id, prop_id, prop_cod, tipo_lic,
                 refresh_cb, refresh_todos_cb):
        super().__init__(parent, bg=C_WHITE,
                         highlightbackground="#DDE3EA", highlightthickness=1)
        self.lic_id = lic_id
        self.prop_id = prop_id
        self.prop_cod = prop_cod
        self.tipo_lic = tipo_lic
        self.refresh_cb = refresh_cb
        self.refresh_todos_cb = refresh_todos_cb
        self.collapsed = True
        self.body_frame = None
        self.solo_falta_var = tk.BooleanVar(value=False)
        self._build_header()

    def _build_header(self):
        conn = get_conn()
        docs = conn.execute(
            "SELECT id, solo_privada, control_web FROM documentos WHERE tipo='propuesta'"
        ).fetchall()
        checks = conn.execute(
            "SELECT estado, web_marcado FROM checklist WHERE licitacion_id=? AND propuesta_id=?",
            (self.lic_id, self.prop_id)
        ).fetchall()
        conn.close()

        completados = 0
        na_count = 0
        for i, (estado, web_marcado) in enumerate(checks):
            if i >= len(docs):
                continue
            doc = docs[i]
            if doc[1] == 1 and self.tipo_lic == "Publica":
                na_count += 1
            elif estado == "NA":
                na_count += 1
            elif estado == "Si":
                requiere_web = doc[2] == 1
                if not requiere_web or web_marcado == 1:
                    completados += 1
        
        total_efectivos = len(docs) - na_count
        pct = int(completados / total_efectivos * 100) if total_efectivos > 0 else 0

        for w in self.winfo_children():
            w.destroy()

        hdr = tk.Frame(self, bg="#EBF5FB")
        hdr.pack(fill=tk.X)

        self.toggle_btn = tk.Button(hdr, text="▶" if self.collapsed else "▼",
                                     font=("Segoe UI", 9), bg="#EBF5FB", fg=C_ACCENT,
                                     relief=tk.FLAT, cursor="hand2",
                                     command=self._toggle)
        self.toggle_btn.pack(side=tk.LEFT, padx=8)

        tk.Label(hdr, text=f"📁  {self.prop_cod}",
                 font=("Segoe UI", 10, "bold"), bg="#EBF5FB", fg=C_TEXT).pack(side=tk.LEFT)

        prog_color = C_SUCCESS if pct == 100 else (C_WARNING if pct >= 50 else C_DANGER)
        self.prog_label = tk.Label(hdr, text=f"  {pct}% ({completados}/{total_efectivos})  ",
                                    font=("Segoe UI", 8, "bold"), bg=prog_color, fg=C_WHITE,
                                    padx=6, pady=2)
        self.prog_label.pack(side=tk.LEFT, padx=10, pady=6)

        tk.Button(hdr, text="✕", font=("Segoe UI", 9), bg="#EBF5FB", fg=C_DANGER,
                  relief=tk.FLAT, cursor="hand2",
                  command=self._eliminar).pack(side=tk.RIGHT, padx=8)

        if not self.collapsed and self.body_frame:
            self._refresh_body_content()

    def _update_progress(self):
        conn = get_conn()
        docs = conn.execute(
            "SELECT id, solo_privada, control_web FROM documentos WHERE tipo='propuesta'"
        ).fetchall()
        checks = conn.execute(
            "SELECT estado, web_marcado FROM checklist WHERE licitacion_id=? AND propuesta_id=?",
            (self.lic_id, self.prop_id)
        ).fetchall()
        conn.close()

        completados = 0
        na_count = 0
        for i, (estado, web_marcado) in enumerate(checks):
            if i >= len(docs):
                continue
            doc = docs[i]
            if doc[1] == 1 and self.tipo_lic == "Publica":
                na_count += 1
            elif estado == "NA":
                na_count += 1
            elif estado == "Si":
                requiere_web = doc[2] == 1
                if not requiere_web or web_marcado == 1:
                    completados += 1
        
        total_efectivos = len(docs) - na_count
        pct = int(completados / total_efectivos * 100) if total_efectivos > 0 else 0

        prog_color = C_SUCCESS if pct == 100 else (C_WARNING if pct >= 50 else C_DANGER)
        self.prog_label.config(text=f"  {pct}% ({completados}/{total_efectivos})  ", bg=prog_color)

    def _toggle(self):
        self.collapsed = not self.collapsed
        self.toggle_btn.config(text="▶" if self.collapsed else "▼")
        if self.collapsed:
            if self.body_frame:
                self.body_frame.destroy()
                self.body_frame = None
        else:
            self._build_body()

    def _build_body(self):
        if self.body_frame:
            self.body_frame.destroy()
        self.body_frame = tk.Frame(self, bg=C_BG)
        self.body_frame.pack(fill=tk.X)

        top = tk.Frame(self.body_frame, bg=C_BG)
        top.pack(fill=tk.X, padx=10, pady=6)
        
        tk.Checkbutton(top, text="Mostrar solo pendientes",
                       variable=self.solo_falta_var,
                       command=self._refresh_body_content,
                       bg=C_BG, fg=C_TEXT, font=("Segoe UI", 8),
                       activebackground=C_BG, selectcolor=C_WHITE).pack(side=tk.LEFT)

        self.docs_frame = tk.Frame(self.body_frame, bg=C_BG)
        self.docs_frame.pack(fill=tk.X, padx=6, pady=(0,8))
        self._refresh_body_content()

    def _refresh_body_content(self):
        if not self.docs_frame:
            return
            
        for w in self.docs_frame.winfo_children():
            w.destroy()

        conn = get_conn()
        docs = conn.execute(
            "SELECT id, nombre, solo_privada, control_web FROM documentos WHERE tipo='propuesta' ORDER BY orden"
        ).fetchall()
        
        checks_raw = conn.execute(
            "SELECT id, estado, documento_id, numero, fecha_doc, web_marcado FROM checklist "
            "WHERE licitacion_id=? AND propuesta_id=?", (self.lic_id, self.prop_id)
        ).fetchall()
        
        checks = {}
        for row in checks_raw:
            chk_id, estado, doc_id, numero, fecha_doc, web_marcado = row
            checks[doc_id] = (chk_id, estado, numero, fecha_doc, web_marcado)
        
        conn.close()

        mostrar_solo_pendientes = self.solo_falta_var.get()
        
        for doc_id, doc_nombre, solo_priv, control_web in docs:
            locked = (solo_priv == 1 and self.tipo_lic == "Publica")
            chk_row = checks.get(doc_id, (None, "No", "", "", 0))
            chk_id, estado, numero, fecha_doc, web_marcado = chk_row
            
            esta_completo = (estado == "Si" and (web_marcado == 1 or control_web == 0))
            es_pendiente = not (estado == "NA" or esta_completo)
            
            if mostrar_solo_pendientes and not es_pendiente:
                continue
            
            DocRow(self.docs_frame, self.lic_id, self.prop_id, doc_id, doc_nombre,
                   estado, numero, fecha_doc, web_marcado,
                   locked, self.tipo_lic, control_web,
                   self._on_doc_update,
                   self.refresh_todos_cb).pack(fill=tk.X, pady=2)

    def _on_doc_update(self):
        self._update_progress()
        self._refresh_body_content()

    def _eliminar(self):
        if not messagebox.askyesno("Confirmar", f"¿Eliminar la propuesta '{self.prop_cod}'?"):
            return
        conn = get_conn()
        conn.execute("DELETE FROM checklist WHERE licitacion_id=? AND propuesta_id=?",
                     (self.lic_id, self.prop_id))
        conn.execute("DELETE FROM propuestas WHERE id=?", (self.prop_id,))
        conn.commit()
        conn.close()
        self.refresh_cb()
        self.refresh_todos_cb()
        self.destroy()


# ═══════════════════════════════════════════════════════════════════════════════
# GENERACIÓN DE PDF
# ═══════════════════════════════════════════════════════════════════════════════
def generar_acta_pdf(path, lic_id, codigo, nombre, tipo, firmante):
     # Calcular ancho automático basado en las columnas de la tabla
    ANCHO_COLUMNAS = [6.5*cm, 2.2*cm, 1.8*cm, 2*cm, 2*cm, 2.2*cm]
    ANCHO_TOTAL = sum(ANCHO_COLUMNAS)
    conn = get_conn()

    docs_lic = conn.execute(
        "SELECT id, nombre, solo_privada, control_web FROM documentos WHERE tipo='licitacion' ORDER BY orden"
    ).fetchall()
    
    checks_lic_raw = conn.execute(
        "SELECT id, estado, documento_id, numero, fecha_doc, web_marcado FROM checklist "
        "WHERE licitacion_id=? AND propuesta_id IS NULL", (lic_id,)
    ).fetchall()
    
    checks_lic = {}
    for row in checks_lic_raw:
        chk_id, estado, doc_id, numero, fecha_doc, web_marcado = row
        checks_lic[doc_id] = (estado, numero, fecha_doc, web_marcado)

    props = conn.execute(
        "SELECT id, codigo FROM propuestas WHERE licitacion_id=? ORDER BY codigo",
        (lic_id,)
    ).fetchall()

    prop_data = []
    for prop_id, prop_cod in props:
        docs_prop = conn.execute(
            "SELECT id, nombre, solo_privada, control_web FROM documentos WHERE tipo='propuesta' ORDER BY orden"
        ).fetchall()
        
        checks_prop_raw = conn.execute(
            "SELECT id, estado, documento_id, numero, fecha_doc, web_marcado FROM checklist "
            "WHERE licitacion_id=? AND propuesta_id=?", (lic_id, prop_id)
        ).fetchall()
        
        checks_prop = {}
        for row in checks_prop_raw:
            chk_id, estado, doc_id, numero, fecha_doc, web_marcado = row
            checks_prop[doc_id] = (estado, numero, fecha_doc, web_marcado)
        
        prop_data.append((prop_cod, docs_prop, checks_prop))
    conn.close()

    doc = SimpleDocTemplate(path, pagesize=A4,
                            leftMargin=2.2*cm, rightMargin=2.2*cm,
                            topMargin=2*cm, bottomMargin=2.5*cm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("title", parent=styles["Title"],
                                  fontSize=16, textColor=colors.HexColor(C_HEADER),
                                  spaceAfter=4, alignment=TA_CENTER)
    subtitle_style = ParagraphStyle("subtitle", parent=styles["Normal"],
                                     fontSize=10, textColor=colors.HexColor(C_MUTED),
                                     alignment=TA_CENTER, spaceAfter=4)
    section_style = ParagraphStyle("section", parent=styles["Heading2"],
                                    fontSize=11, textColor=colors.white,
                                    leading=20, spaceBefore=14, spaceAfter=6,
                                    leftIndent=0, rightIndent=0,
                                    borderPadding=(4, 8, 4, 8))
    prop_style = ParagraphStyle("prop", parent=styles["Heading3"],
                                  fontSize=10, textColor=colors.HexColor(C_ACCENT),
                                  spaceBefore=10, spaceAfter=4,
                                  borderPadding=(4, 8, 4, 8))
    normal_style = ParagraphStyle("normal", parent=styles["Normal"],
                                   fontSize=9, textColor=colors.HexColor("#2C3E50"),
    # Texto más oscuro
                                   leading=14)  # Más espacio entre líneas

    # Estilo para encabezados de tabla
    header_style = ParagraphStyle("header", parent=styles["Normal"],
                                   fontSize=9, textColor=colors.white,
                                   fontName="Helvetica-Bold", leading=12)


    story = []

    # Título con marco
    story.append(Spacer(1, 10))
    title_frame = Table([[Paragraph("ACTA DE CIERRE DE CARPETA", title_style)]], colWidths=[16*cm])
    title_frame.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#EBF5FB")),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("PADDING", (0,0), (-1,-1), 15),
        ("GRID", (0,0), (-1,-1), 1, colors.HexColor(C_ACCENT)),
    ]))
    story.append(title_frame)
    story.append(Spacer(1, 5))
    story.append(Paragraph("Sistema de Gestión Documental de Licitaciones", subtitle_style))
    story.append(Spacer(1, 5))
    story.append(HRFlowable(width="80%", thickness=1, color=colors.HexColor(C_ACCENT), hAlign="CENTER"))
    story.append(Spacer(1, 14))

    tipo_color = colors.HexColor(C_DANGER if tipo == "Privada" else C_SUCCESS)
    info_data = [
        ["Código:", codigo, "Tipo:", tipo],
        ["Nombre:", Paragraph(nombre, normal_style), "Fecha:", datetime.now().strftime("%d/%m/%Y")],
    ]
    info_table = Table(info_data, colWidths=[3*cm, 7*cm, 2.5*cm, 3.5*cm])
    info_table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTNAME", (2,0), (2,-1), "Helvetica-Bold"),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E8F4FD")),
        ("BACKGROUND", (0,1), (-1,1), colors.HexColor("#F0F8FF")),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#BDC3C7")),
        ("PADDING", (0,0), (-1,-1), 8),
        ("TEXTCOLOR", (3,0), (3,0), tipo_color),
        ("FONTNAME", (3,0), (3,0), "Helvetica-Bold"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 16))

    def tabla_docs(docs, checks, tipo_lic):
        ESTADO_LABEL = {"Si": "✔  Sí", "No": "✗  No", "NA": "—  N/A"}
        ESTADO_COLOR = {
            "Si": colors.HexColor(C_SUCCESS),
            "No": colors.HexColor(C_DANGER),
            "NA": colors.HexColor(C_NA),
        }

        # Encabezados con mejor formato
        hdr = [
            Paragraph("<b>Documento</b>", header_style),
            Paragraph("<b>Estado Doc.</b>", header_style),
            Paragraph("<b>Web</b>", header_style),
            Paragraph("<b>Completo</b>", header_style),
            Paragraph("<b>N°</b>", header_style),
            Paragraph("<b>Fecha</b>", header_style),
        ]
        rows = [hdr]
        
        for doc_id, doc_nombre, solo_priv, control_web in docs:
            locked = (solo_priv == 1 and tipo_lic == "Publica")
            chk = checks.get(doc_id, ("No", "", "", 0))
            estado, numero, fecha_doc, web_marcado = chk
            if locked:
                estado = "NA"
                web_marcado = 0
            nl = nombre_limpio(doc_nombre)
            if solo_priv:
                nl += " *"
            if control_web == 1:
                nl += " 🌐"
            
            esta_completo = (estado == "Si" and (web_marcado == 1 or control_web == 0))
            completo_texto = "✅ Sí" if esta_completo else "❌ No"
            if estado == "NA":
                completo_texto = "— N/A"
            
            # Crear celdas con formato mejorado
            celda_doc = Paragraph(nl, normal_style)
            
            # Color para el estado del documento
            estado_color = ESTADO_COLOR.get(estado, colors.black)
            celda_estado = Paragraph(f'<font color="{estado_color.hexval()}">{ESTADO_LABEL.get(estado, estado)}</font>', normal_style)
            
            # Estado web
            if control_web == 1:
                web_texto = "🌐 Sí" if web_marcado == 1 else "◻ No"
                web_color = colors.HexColor(C_WEB_OK) if web_marcado == 1 else colors.HexColor(C_MUTED)
                celda_web = Paragraph(f'<font color="{web_color.hexval()}">{web_texto}</font>', normal_style)
            else:
                celda_web = Paragraph("—", normal_style)
            
            # Columna completo
            completo_color = colors.HexColor(C_SUCCESS) if esta_completo else (colors.HexColor(C_DANGER) if estado != "NA" else colors.HexColor(C_NA))
            celda_completo = Paragraph(f'<font color="{completo_color.hexval()}"><b>{completo_texto}</b></font>', normal_style)
            
            rows.append([
                celda_doc,
                celda_estado,
                celda_web,
                celda_completo,
                Paragraph(numero or "—", normal_style),
                Paragraph(fecha_doc or "—", normal_style),
            ])

        # Ajustar anchos de columna para mejor distribución
        t = Table(rows, colWidths=[6.5*cm, 2.2*cm, 1.8*cm, 2*cm, 2*cm, 2.2*cm])
        ts = TableStyle([
            # Encabezado
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor(C_HEADER)),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,0), 9),
            ("ALIGN", (0,0), (-1,0), "CENTER"),
            ("VALIGN", (0,0), (-1,0), "MIDDLE"),
            ("BOTTOMPADDING", (0,0), (-1,0), 8),
            ("TOPPADDING", (0,0), (-1,0), 8),
            
            # Cuerpo de la tabla
            ("FONTSIZE", (0,1), (-1,-1), 8),
            ("ALIGN", (1,1), (1,-1), "CENTER"),  # Estado centrado
            ("ALIGN", (2,1), (2,-1), "CENTER"),  # Web centrado
            ("ALIGN", (3,1), (3,-1), "CENTER"),  # Completo centrado
            ("ALIGN", (4,1), (4,-1), "CENTER"),  # N° centrado
            ("ALIGN", (5,1), (5,-1), "CENTER"),  # Fecha centrado
            ("VALIGN", (0,1), (-1,-1), "TOP"),
            
            # Bordes y colores
            ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#BDC3C7")),
            ("ROWBACKGROUNDS", (0,1), (-1,-1),
            [colors.white, colors.HexColor("#F8F9FA")]),
            ("PADDING", (0,0), (-1,-1), 6),
            
            # Borde superior más grueso para el encabezado
            ("LINEBELOW", (0,0), (-1,0), 0.5, colors.HexColor("#BDC3C7")),
        ])
        t.setStyle(ts)
        return t

    # Título de sección con borde negro y fondo azul (sin duplicar)
    section_title = Table([[Paragraph("DOCUMENTACIÓN DE LA LICITACIÓN", section_style)]], 
                        colWidths=ANCHO_TOTAL)
    section_title.setStyle(TableStyle([
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor(C_HEADER)),  # Único fondo
        ("LINEBELOW", (0,0), (-1,-1), 0.5, colors.HexColor("#BDC3C7")),
        ("PADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(section_title)
    story.append(Spacer(1, 8))
    story.append(tabla_docs(docs_lic, checks_lic, tipo))
    if tipo == "Publica":
        story.append(Paragraph("* Documento aplica solo a licitaciones privadas (marcado N/A automáticamente).",
                                ParagraphStyle("nota", parent=styles["Normal"],
                                               fontSize=7, textColor=colors.HexColor(C_MUTED),
                                               spaceBefore=2)))
    story.append(Spacer(1, 12))

    if prop_data:
        section_title_prop = Table([[Paragraph("DOCUMENTACIÓN POR PROPUESTA", section_style)]], 
                            colWidths=ANCHO_TOTAL)
        section_title_prop.setStyle(TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("BACKGROUND", (0,0), (-1,-1), colors.HexColor(C_HEADER)),
            ("GRID", (0,0), (-1,-1), 1, colors.black),
            ("PADDING", (0,0), (-1,-1), 6),
        ]))
        story.append(section_title_prop)                    
        for prop_cod, docs_prop, checks_prop in prop_data:
            story.append(Spacer(1, 6))
            story.append(Paragraph(f"  Propuesta: {prop_cod}", prop_style))
            story.append(tabla_docs(docs_prop, checks_prop, tipo))

    story.append(Spacer(1, 30))

    # Espacio antes de la firma
    story.append(Spacer(1, 40))

    # Línea de firma más elegante
    story.append(HRFlowable(width="40%", thickness=1, color=colors.HexColor(C_ACCENT),
                            hAlign="CENTER"))
    story.append(Spacer(1, 10))

    # Firma
    firma_style = ParagraphStyle("firma_nombre", parent=styles["Normal"],
                                fontSize=11, textColor=colors.HexColor(C_HEADER),
                                fontName="Helvetica-Bold", alignment=TA_CENTER,
                                spaceAfter=4)
    story.append(Paragraph(firmante, firma_style))

    cargo_style = ParagraphStyle("firma_cargo", parent=styles["Normal"],
                                fontSize=9, textColor=colors.HexColor(C_MUTED),
                                alignment=TA_CENTER)
    story.append(Paragraph("Encargado de Carpeta de Licitación", cargo_style))

    story.append(Spacer(1, 15))

    # Fecha de emisión
    fecha_style = ParagraphStyle("fecha_emision", parent=styles["Normal"],
                                fontSize=8, textColor=colors.HexColor(C_MUTED),
                                alignment=TA_CENTER)
    story.append(Paragraph(f"Documento emitido el {datetime.now().strftime('%d/%m/%Y')}", fecha_style))
        

    doc.build(story)


# ═══════════════════════════════════════════════════════════════════════════════
# WIDGETS AUXILIARES
# ═══════════════════════════════════════════════════════════════════════════════
def scrollable(parent):
    outer = tk.Frame(parent, bg=C_BG)
    outer.pack(fill=tk.BOTH, expand=True)
    canvas = tk.Canvas(outer, bg=C_BG, highlightthickness=0)
    sb = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=canvas.yview)
    inner = tk.Frame(canvas, bg=C_BG)
    inner_id = canvas.create_window((0,0), window=inner, anchor="nw")

    def on_configure(e):
        canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.itemconfig(inner_id, width=canvas.winfo_width())

    inner.bind("<Configure>", on_configure)
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(inner_id, width=e.width))
    canvas.configure(yscrollcommand=sb.set)

    def mousewheel(e):
        canvas.yview_scroll(int(-1*(e.delta/120)), "units")
    
    def on_enter(e):
        canvas.bind_all("<MouseWheel>", mousewheel)
    
    def on_leave(e):
        canvas.unbind_all("<MouseWheel>")
    
    canvas.bind("<Enter>", on_enter)
    canvas.bind("<Leave>", on_leave)

    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    sb.pack(side=tk.RIGHT, fill=tk.Y)
    return canvas, inner


class HrLine(tk.Frame):
    def __init__(self, parent, color):
        super().__init__(parent, bg=color, height=1)


class SideBtn(tk.Button):
    def __init__(self, parent, text, command):
        super().__init__(parent, text=text, command=command,
                         bg="#2C5282", fg=C_WHITE,
                         font=("Segoe UI", 9), relief=tk.FLAT,
                         cursor="hand2", pady=7, anchor="w", padx=10)
        self.bind("<Enter>", lambda _: self.config(bg=C_ACCENT))
        self.bind("<Leave>", lambda _: self.config(bg="#2C5282"))


class HdrBtn(tk.Button):
    def __init__(self, parent, text, command):
        super().__init__(parent, text=text, command=command,
                         bg=C_ACCENT, fg=C_WHITE,
                         font=("Segoe UI", 9), relief=tk.FLAT,
                         cursor="hand2", pady=5, padx=12)
        self.bind("<Enter>", lambda _: self.config(bg="#1A6FA8"))
        self.bind("<Leave>", lambda _: self.config(bg=C_ACCENT))


class SectionHeader(tk.Frame):
    def __init__(self, parent, text):
        super().__init__(parent, bg="#EBF5FB")
        tk.Label(self, text=text, font=("Segoe UI", 10, "bold"),
                 bg="#EBF5FB", fg=C_HEADER, pady=8, padx=12,
                 anchor="w").pack(fill=tk.X)


def simple_input(parent, prompt):
    result = [None]
    dlg = tk.Toplevel(parent)
    dlg.title(prompt)
    dlg.geometry("320x140")
    dlg.configure(bg=C_BG)
    dlg.grab_set()
    dlg.resizable(False, False)
    tk.Label(dlg, text=prompt + ":", font=("Segoe UI", 10),
             bg=C_BG, fg=C_TEXT).pack(pady=(18,6))
    e = tk.Entry(dlg, font=("Segoe UI", 11), relief=tk.SOLID, bd=1)
    e.pack(padx=30, fill=tk.X, ipady=5)
    e.focus()
    def ok():
        result[0] = e.get().strip()
        dlg.destroy()
    tk.Button(dlg, text="Aceptar", command=ok,
              bg=C_ACCENT, fg=C_WHITE, font=("Segoe UI", 10),
              relief=tk.FLAT, cursor="hand2", pady=6).pack(pady=14, padx=30, fill=tk.X)
    dlg.bind("<Return>", lambda _: ok())
    dlg.wait_window()
    return result[0]


if __name__ == "__main__":
    app = App()
    app.mainloop()